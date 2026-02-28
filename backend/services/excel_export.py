"""
Excel report generation (SPEC.md "Output: Excel File" section).

Creates a 3-tab .xlsx file:
  Tab 1: "Creator Payout Summary" — one row per creator (from CreatorSummary)
  Tab 2: "Video Audit"           — one row per payout unit (from PayoutUnit)
  Tab 3: "Exceptions"            — flagged videos for review (from ExceptionVideo)

File naming: "Polymarket Payout Summary {start_date} to {end_date}.xlsx"

Formatting:
  - Bold header rows on all tabs
  - Auto-fit column widths (with min/max constraints)
  - Freeze top row (header) on all tabs
  - Currency format for payout columns ($#,##0.00)
  - Comma-separated number format for view counts (#,##0)
  - Sorted per SPEC.md requirements
"""

import os
import logging
from datetime import date, datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from models.schemas import CreatorSummary, PayoutUnit, ExceptionVideo

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MIN_COL_WIDTH = 10      # Minimum column width (characters)
MAX_COL_WIDTH = 50      # Maximum column width (avoid super-wide columns)
HEADER_FONT = Font(bold=True)
CURRENCY_FORMAT = '$#,##0.00'
NUMBER_FORMAT = '#,##0'


# ===========================================================================
# Public API
# ===========================================================================

def generate_report(
    summaries: list[CreatorSummary],
    payout_units: list[PayoutUnit],
    exceptions: list[ExceptionVideo],
    start_date: date,
    end_date: date,
    output_dir: Optional[str] = None,
) -> str:
    """
    Generate the .xlsx payout report with 3 tabs.

    Args:
        summaries:    Per-creator summary rows for Tab 1
        payout_units: Per-video payout rows for Tab 2
        exceptions:   Exception videos for Tab 3
        start_date:   Payout period start (for filename)
        end_date:     Payout period end (for filename)
        output_dir:   Directory to save the file (defaults to config.OUTPUT_DIR)

    Returns:
        Absolute file path of the generated .xlsx report.
    """
    if output_dir is None:
        output_dir = config.OUTPUT_DIR

    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # ------------------------------------------------------------------
    # Build filename
    # ------------------------------------------------------------------
    filename = (
        f"Polymarket Payout Summary "
        f"{start_date.isoformat()} to {end_date.isoformat()}.xlsx"
    )
    filepath = os.path.join(output_dir, filename)

    logger.info(f"Generating report: {filepath}")

    # ------------------------------------------------------------------
    # Create workbook and tabs
    # ------------------------------------------------------------------
    wb = Workbook()

    # Tab 1: Creator Payout Summary (default sheet, rename it)
    ws1 = wb.active
    ws1.title = "Creator Payout Summary"
    _build_tab1_creator_summary(ws1, summaries)

    # Tab 2: Video Audit
    ws2 = wb.create_sheet("Video Audit")
    _build_tab2_video_audit(ws2, payout_units)

    # Tab 3: Exceptions
    ws3 = wb.create_sheet("Exceptions")
    _build_tab3_exceptions(ws3, exceptions)

    # ------------------------------------------------------------------
    # Save
    # ------------------------------------------------------------------
    wb.save(filepath)
    logger.info(
        f"Report saved: {filepath} "
        f"({len(summaries)} creators, {len(payout_units)} payout units, "
        f"{len(exceptions)} exceptions)"
    )

    return filepath


# ===========================================================================
# Tab 1: Creator Payout Summary
# ===========================================================================

def _build_tab1_creator_summary(
    ws: Worksheet,
    summaries: list[CreatorSummary],
) -> None:
    """
    Tab 1: One row per creator with aggregated payout data.

    Columns:
      Creator Name | Qualified Video Count | Total Payout |
      Paired Video Count | Exception Count

    Sorted by Total Payout descending.
    """
    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    headers = [
        "Creator Name",
        "Qualified Video Count",
        "Total Payout",
        "Paired Video Count",
        "Exception Count",
    ]
    ws.append(headers)

    # ------------------------------------------------------------------
    # Data rows — sorted by total_payout descending
    # ------------------------------------------------------------------
    sorted_summaries = sorted(summaries, key=lambda s: s.total_payout, reverse=True)

    for s in sorted_summaries:
        ws.append([
            s.creator_name,
            s.qualified_video_count,
            s.total_payout,
            s.paired_video_count,
            s.exception_count,
        ])

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------
    _format_header_row(ws)
    _freeze_top_row(ws)

    # Currency format for Total Payout column (column C = index 3)
    _apply_column_format(ws, col_idx=3, fmt=CURRENCY_FORMAT, start_row=2)

    # Number format for count columns (columns B, D, E)
    for col_idx in [2, 4, 5]:
        _apply_column_format(ws, col_idx=col_idx, fmt=NUMBER_FORMAT, start_row=2)

    _auto_fit_columns(ws)


# ===========================================================================
# Tab 2: Video Audit
# ===========================================================================

def _build_tab2_video_audit(
    ws: Worksheet,
    payout_units: list[PayoutUnit],
) -> None:
    """
    Tab 2: One row per payout unit — only paired videos appear here.

    Columns:
      Creator Name | Uploaded At | Video Length (sec) |
      TikTok Link | TikTok Views | Instagram Link | Instagram Views |
      Chosen Views | Effective Views | Payout Amount |
      Match Method | Match Notes | Latest Updated At

    Sorted by Creator Name, then Uploaded At.
    """
    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    headers = [
        "Creator Name",
        "Uploaded At",
        "Video Length (sec)",
        "TikTok Link",
        "TikTok Views",
        "Instagram Link",
        "Instagram Views",
        "Chosen Views",
        "Effective Views",
        "Payout Amount",
        "Match Method",
        "Match Notes",
        "Latest Updated At",
    ]
    ws.append(headers)

    # ------------------------------------------------------------------
    # Data rows — sorted by Creator Name, then Uploaded At
    # ------------------------------------------------------------------
    sorted_units = sorted(payout_units, key=_tab2_sort_key)

    for pu in sorted_units:
        # Both videos are always present (only paired units reach Tab 2)
        tt_link = pu.tiktok_video.ad_link
        tt_views = pu.tiktok_video.latest_views
        ig_link = pu.instagram_video.ad_link
        ig_views = pu.instagram_video.latest_views

        uploaded_at = _get_uploaded_at(pu)
        video_length = _get_video_length(pu)
        latest_updated = _get_latest_updated_at(pu)

        ws.append([
            pu.creator_name,
            _format_date(uploaded_at),
            video_length,
            tt_link,
            tt_views,
            ig_link,
            ig_views,
            pu.chosen_views,
            pu.effective_views,
            pu.payout_amount,
            pu.match_method,
            pu.match_note,
            _format_datetime(latest_updated),
        ])

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------
    _format_header_row(ws)
    _freeze_top_row(ws)

    # Views columns with comma separators (E=5, G=7, H=8, I=9)
    for col_idx in [5, 7, 8, 9]:
        _apply_column_format(ws, col_idx=col_idx, fmt=NUMBER_FORMAT, start_row=2)

    # Currency format for Payout Amount (column J = 10)
    _apply_column_format(ws, col_idx=10, fmt=CURRENCY_FORMAT, start_row=2)

    _auto_fit_columns(ws)


# ===========================================================================
# Tab 3: Exceptions
# ===========================================================================

def _build_tab3_exceptions(
    ws: Worksheet,
    exceptions: list[ExceptionVideo],
) -> None:
    """
    Tab 3: Exception videos flagged for manual review.

    Columns:
      Username | Platform | Video Link | Created At |
      Latest Views | Video Length (sec) | Reason

    Includes ALL exceptions: mapping failures, private/removed, unpaired, etc.
    """
    # ------------------------------------------------------------------
    # Header row
    # ------------------------------------------------------------------
    headers = [
        "Username",
        "Platform",
        "Video Link",
        "Created At",
        "Latest Views",
        "Video Length (sec)",
        "Reason",
    ]
    ws.append(headers)

    # ------------------------------------------------------------------
    # Data rows
    # ------------------------------------------------------------------
    for exc in exceptions:
        ws.append([
            exc.username,
            exc.platform,
            exc.ad_link,
            _format_datetime(exc.created_at),
            exc.latest_views,
            exc.video_length,
            exc.reason,
        ])

    # ------------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------------
    _format_header_row(ws)
    _freeze_top_row(ws)

    # Views column with comma separators (column E = 5)
    _apply_column_format(ws, col_idx=5, fmt=NUMBER_FORMAT, start_row=2)

    _auto_fit_columns(ws)


# ===========================================================================
# Formatting helpers
# ===========================================================================

def _format_header_row(ws: Worksheet) -> None:
    """Bold the entire header row (row 1)."""
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _freeze_top_row(ws: Worksheet) -> None:
    """Freeze the top row so the header stays visible when scrolling."""
    ws.freeze_panes = "A2"


def _apply_column_format(
    ws: Worksheet,
    col_idx: int,
    fmt: str,
    start_row: int = 2,
) -> None:
    """
    Apply a number format to all data cells in a column.

    Args:
        ws:        Worksheet
        col_idx:   1-based column index
        fmt:       Number format string (e.g., '$#,##0.00' or '#,##0')
        start_row: First data row (skip header)
    """
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            cell.number_format = fmt


def _auto_fit_columns(ws: Worksheet) -> None:
    """
    Auto-fit column widths based on cell content.

    Examines header + all data rows to find the widest value in each column,
    then sets the column width with min/max constraints.
    """
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)

        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None:
                # Estimate width from string representation
                cell_text = str(cell.value)
                cell_length = len(cell_text)
                if cell_length > max_length:
                    max_length = cell_length

        # Apply width with padding and constraints
        # Add 2 chars of padding for readability
        adjusted_width = max_length + 2
        adjusted_width = max(adjusted_width, MIN_COL_WIDTH)
        adjusted_width = min(adjusted_width, MAX_COL_WIDTH)
        ws.column_dimensions[col_letter].width = adjusted_width


# ===========================================================================
# Data extraction helpers (for Tab 2)
# ===========================================================================

def _get_uploaded_at(pu: PayoutUnit) -> Optional[date]:
    """Get uploaded_at from the PayoutUnit (prefer TikTok, fallback to Instagram)."""
    if pu.tiktok_video and pu.tiktok_video.uploaded_at:
        return pu.tiktok_video.uploaded_at
    if pu.instagram_video and pu.instagram_video.uploaded_at:
        return pu.instagram_video.uploaded_at
    return None


def _get_video_length(pu: PayoutUnit) -> Optional[int]:
    """Get video_length from the PayoutUnit (prefer TikTok, fallback to Instagram)."""
    if pu.tiktok_video and pu.tiktok_video.video_length is not None:
        return pu.tiktok_video.video_length
    if pu.instagram_video and pu.instagram_video.video_length is not None:
        return pu.instagram_video.video_length
    return None


def _get_latest_updated_at(pu: PayoutUnit) -> Optional[datetime]:
    """Get the most recent latest_updated_at from either video."""
    tt_updated = pu.tiktok_video.latest_updated_at if pu.tiktok_video else None
    ig_updated = pu.instagram_video.latest_updated_at if pu.instagram_video else None

    if tt_updated and ig_updated:
        return max(tt_updated, ig_updated)
    return tt_updated or ig_updated


def _tab2_sort_key(pu: PayoutUnit) -> tuple:
    """
    Sort key for Tab 2: Creator Name ascending, then Uploaded At ascending.
    None values sort to the end.
    """
    uploaded_at = _get_uploaded_at(pu)
    return (
        pu.creator_name or "",
        uploaded_at or date.max,
    )


# ===========================================================================
# Date formatting helpers
# ===========================================================================

def _format_date(d: Optional[date]) -> Optional[str]:
    """Format a date as YYYY-MM-DD, or None if missing."""
    if d is None:
        return None
    return d.isoformat()


def _format_datetime(dt: Optional[datetime]) -> Optional[str]:
    """Format a datetime as YYYY-MM-DD HH:MM:SS, or None if missing."""
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M:%S")


# ===========================================================================
# Standalone test — run with: cd backend && python -m services.excel_export
# ===========================================================================

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")

    # Create synthetic test data
    summaries = [
        CreatorSummary(
            creator_name="Alice", qualified_video_count=2,
            total_payout=200.0, paired_video_count=2,
            exception_count=0,
        ),
        CreatorSummary(
            creator_name="Bob", qualified_video_count=1,
            total_payout=100.0, paired_video_count=1,
            exception_count=1,
        ),
    ]

    from models.schemas import Video
    tt_video = Video(
        username="alice_tt", platform="tiktok",
        ad_link="https://tiktok.com/@alice/video/123",
        uploaded_at=date(2026, 2, 20),
        created_at=datetime(2026, 2, 20, 10, 0, 0),
        video_length=30, latest_views=50000,
        latest_updated_at=datetime(2026, 2, 21, 12, 0, 0),
    )
    ig_video = Video(
        username="alice_ig", platform="instagram",
        ad_link="https://instagram.com/p/abc123",
        uploaded_at=date(2026, 2, 20),
        created_at=datetime(2026, 2, 20, 10, 30, 0),
        video_length=30, latest_views=80000,
        latest_updated_at=datetime(2026, 2, 21, 14, 0, 0),
    )

    payout_units = [
        PayoutUnit(
            creator_name="Alice", tiktok_video=tt_video, instagram_video=ig_video,
            chosen_views=80000, effective_views=80000, best_platform="instagram",
            payout_amount=100.0, match_method="sequence",
            match_note="sequence match, phash distance: 0", phash_distance=0,
        ),
    ]

    exceptions = [
        ExceptionVideo(
            username="unknown_user", platform="tiktok",
            ad_link="https://tiktok.com/@unknown/video/456",
            created_at=datetime(2026, 2, 20, 15, 0, 0),
            latest_views=1000, video_length=45,
            reason="not in creator list",
        ),
    ]

    filepath = generate_report(
        summaries, payout_units, exceptions,
        start_date=date(2026, 2, 20), end_date=date(2026, 2, 21),
        output_dir="/tmp/payout_test",
    )
    print(f"\nReport generated: {filepath}")
