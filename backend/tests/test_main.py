"""
Comprehensive tests for main.py — FastAPI endpoints and pipeline wiring.

Tests verify:
  1. POST /api/calculate — full pipeline (with mocked external services)
     - Happy path: valid dates, videos fetched, report generated
     - Date validation: start > end → 400
     - Single-day period: start == end → 200
     - No videos found → 200 with empty report
     - Shortimize API failure → 502
     - Google Sheet failure → 502
  2. GET /api/download/{filename}
     - Valid file → 200 with correct headers
     - Missing file → 404
  3. Response format verification
     - JSON structure matches spec
     - Summary fields present and correct types
  4. Exception count aggregation
     - _count_exceptions_per_creator correctness
     - Unmappable exceptions don't inflate counts
     - Mixed exception sources
  5. Pipeline integration
     - Multi-creator scenarios
     - Exception counts wired correctly to CreatorSummary

All external dependencies (Shortimize API, Google Sheets) are mocked
so tests run fast and don't require network access.
"""

import sys
import os
import pytest
import tempfile
import shutil
from datetime import date, datetime
from unittest.mock import patch, MagicMock

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from fastapi.testclient import TestClient

from models.schemas import Video, ExceptionVideo, Creator
from main import app, _count_exceptions_per_creator


# ===========================================================================
# Fixtures
# ===========================================================================

@pytest.fixture
def client():
    """FastAPI test client."""
    return TestClient(app)


@pytest.fixture
def output_dir(tmp_path):
    """Override OUTPUT_DIR with a temp directory."""
    with patch("main.config.OUTPUT_DIR", str(tmp_path)):
        with patch("services.excel_export.config.OUTPUT_DIR", str(tmp_path)):
            os.makedirs(str(tmp_path), exist_ok=True)
            yield str(tmp_path)


# ===========================================================================
# Test data helpers
# ===========================================================================

def make_video(
    username="alice_tt", platform="tiktok", length=30, views=5000,
    created_at_str="2026-02-20T10:00:00+00:00", ad_link=None,
):
    return Video(
        username=username, platform=platform,
        ad_link=ad_link or f"https://{platform}.com/@{username}/video/123",
        uploaded_at=date(2026, 2, 20),
        created_at=datetime.fromisoformat(created_at_str),
        video_length=length, latest_views=views,
        latest_updated_at=datetime.fromisoformat(created_at_str),
    )


def make_exception(username="unknown", platform="tiktok", reason="not in creator list"):
    return ExceptionVideo(
        username=username, platform=platform,
        ad_link=f"https://{platform}.com/@{username}/456",
        created_at=datetime(2026, 2, 20, 15, 0, 0),
        latest_views=1000, video_length=45, reason=reason,
    )


MOCK_CREATORS = [
    Creator(creator_name="Alice", tiktok_handle="alice_tt", instagram_handle="alice_ig"),
    Creator(creator_name="Bob", tiktok_handle="bob_tt", instagram_handle="bob_ig"),
]
MOCK_TT_MAP = {"alice_tt": "Alice", "bob_tt": "Bob"}
MOCK_IG_MAP = {"alice_ig": "Alice", "bob_ig": "Bob"}

MOCK_VIDEOS = [
    # Alice: 1 TT + 1 IG (same length → will pair)
    make_video("alice_tt", "tiktok", 30, 50000, "2026-02-20T10:00:00+00:00", "tt_alice_1"),
    make_video("alice_ig", "instagram", 30, 80000, "2026-02-20T10:30:00+00:00", "ig_alice_1"),
    # Bob: 1 TT only (will be unpaired)
    make_video("bob_tt", "tiktok", 45, 5000, "2026-02-20T14:00:00+00:00", "tt_bob_1"),
]

MOCK_API_EXCEPTIONS = [
    make_exception("private_user", "tiktok", "video marked private"),
]


# ===========================================================================
# 1. POST /api/calculate — Happy path
# ===========================================================================

class TestCalculateHappyPath:
    """Full pipeline with mocked external services."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_success_response(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, MOCK_API_EXCEPTIONS)

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "success"
        assert data["filename"].startswith("Polymarket Payout Summary")
        assert data["filename"].endswith(".xlsx")

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_summary_fields_present(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, MOCK_API_EXCEPTIONS)

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        data = response.json()
        summary = data["summary"]
        assert "total_creators" in summary
        assert "total_payout" in summary
        assert "total_videos_processed" in summary
        assert "total_paired" in summary
        assert "total_unpaired" in summary
        assert "total_exceptions" in summary

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_summary_values_correct(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """
        With our mock data:
          Alice: 1 pair (TT 50K + IG 80K → chosen=80K → $100)
          Bob: 1 unpaired TT (5K → $35)
          api_exceptions: 1 (private_user)
          match_exceptions: 1 (Bob unpaired)
        """
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, MOCK_API_EXCEPTIONS)

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        data = response.json()
        summary = data["summary"]
        assert summary["total_creators"] == 2       # Alice + Bob
        assert summary["total_payout"] == 135.0      # $100 + $35
        assert summary["total_videos_processed"] == 3 # 3 valid videos in
        assert summary["total_paired"] == 1           # Alice's pair
        assert summary["total_unpaired"] == 1         # Bob's standalone
        # Exceptions: 1 api (private) + 1 match (Bob unpaired)
        assert summary["total_exceptions"] == 2

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_xlsx_file_created(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, [])

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        data = response.json()
        filepath = os.path.join(output_dir, data["filename"])
        assert os.path.exists(filepath)

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_filename_contains_dates(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, [])

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        data = response.json()
        assert "2026-02-20" in data["filename"]
        assert "2026-02-21" in data["filename"]


# ===========================================================================
# 2. POST /api/calculate — Date validation
# ===========================================================================

class TestCalculateDateValidation:
    """Date validation edge cases."""

    def test_start_after_end_returns_400(self, client):
        response = client.post("/api/calculate", json={
            "start_date": "2026-02-22",
            "end_date": "2026-02-20",
        })
        assert response.status_code == 400

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_same_day_period(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """start_date == end_date → valid single-day period."""
        mock_fetch_mapping.return_value = ([], {}, {})
        mock_fetch_videos.return_value = ([], [])

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-20",
        })
        assert response.status_code == 200

    def test_invalid_date_format(self, client):
        """Non-date string → 422 validation error."""
        response = client.post("/api/calculate", json={
            "start_date": "not-a-date",
            "end_date": "2026-02-21",
        })
        assert response.status_code == 422


# ===========================================================================
# 3. POST /api/calculate — Error handling
# ===========================================================================

class TestCalculateErrorHandling:
    """Error handling for external service failures."""

    @patch("main.fetch_creator_mapping")
    def test_google_sheet_failure_returns_502(self, mock_fetch_mapping, client):
        mock_fetch_mapping.side_effect = Exception("Network error")

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })
        assert response.status_code == 502
        data = response.json()
        assert "creator mapping" in data["detail"]["message"].lower()

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_shortimize_failure_returns_502(self, mock_fetch_videos, mock_fetch_mapping, client):
        mock_fetch_mapping.return_value = ([], {}, {})
        mock_fetch_videos.side_effect = RuntimeError("API 500 error")

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })
        assert response.status_code == 502
        data = response.json()
        assert "shortimize" in data["detail"]["message"].lower() or \
               "video data" in data["detail"]["message"].lower()

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_no_videos_returns_empty_report(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """No videos found → 200 with empty report."""
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = ([], [])  # No videos

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })
        assert response.status_code == 200
        data = response.json()
        assert data["status"] == "success"
        assert data["summary"]["total_creators"] == 0
        assert data["summary"]["total_payout"] == 0
        assert data["summary"]["total_videos_processed"] == 0


# ===========================================================================
# 4. GET /api/download/{filename}
# ===========================================================================

class TestDownloadEndpoint:
    """Test the file download endpoint."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_download_existing_file(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """Generate a report, then download it."""
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, [])

        # Generate report
        calc_response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })
        filename = calc_response.json()["filename"]

        # Download it
        response = client.get(f"/api/download/{filename}")
        assert response.status_code == 200
        assert "attachment" in response.headers.get("content-disposition", "")
        assert response.headers.get("content-type") == \
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_download_nonexistent_file_returns_404(self, client):
        response = client.get("/api/download/nonexistent.xlsx")
        assert response.status_code == 404


# ===========================================================================
# 5. _count_exceptions_per_creator tests
# ===========================================================================

class TestCountExceptionsPerCreator:
    """Test the exception count aggregation helper."""

    def test_mapped_exceptions_counted(self):
        """Exceptions for known creators are counted."""
        exceptions = [
            make_exception("alice_tt", "tiktok", "video marked private"),
            make_exception("alice_ig", "instagram", "unpaired — single platform only"),
            make_exception("bob_tt", "tiktok", "video removed"),
        ]
        counts = _count_exceptions_per_creator(exceptions, MOCK_TT_MAP, MOCK_IG_MAP)
        assert counts["Alice"] == 2  # alice_tt + alice_ig
        assert counts["Bob"] == 1    # bob_tt

    def test_unmapped_exceptions_skipped(self):
        """Exceptions for unknown users don't appear in counts."""
        exceptions = [
            make_exception("unknown_user", "tiktok", "not in creator list"),
        ]
        counts = _count_exceptions_per_creator(exceptions, MOCK_TT_MAP, MOCK_IG_MAP)
        assert len(counts) == 0

    def test_mixed_mapped_and_unmapped(self):
        exceptions = [
            make_exception("alice_tt", "tiktok", "unpaired — single platform only"),
            make_exception("stranger", "instagram", "not in creator list"),
            make_exception("bob_ig", "instagram", "video marked private"),
        ]
        counts = _count_exceptions_per_creator(exceptions, MOCK_TT_MAP, MOCK_IG_MAP)
        assert counts.get("Alice") == 1
        assert counts.get("Bob") == 1
        assert "stranger" not in counts

    def test_empty_exceptions(self):
        counts = _count_exceptions_per_creator([], MOCK_TT_MAP, MOCK_IG_MAP)
        assert counts == {}

    def test_case_insensitive_lookup(self):
        """Username lookup should be case-insensitive."""
        exceptions = [
            make_exception("Alice_TT", "tiktok", "some reason"),
        ]
        counts = _count_exceptions_per_creator(exceptions, MOCK_TT_MAP, MOCK_IG_MAP)
        assert counts.get("Alice") == 1

    def test_multiple_exceptions_same_creator(self):
        """Multiple exceptions for one creator are summed."""
        exceptions = [
            make_exception("alice_tt", "tiktok", "reason 1"),
            make_exception("alice_tt", "tiktok", "reason 2"),
            make_exception("alice_ig", "instagram", "reason 3"),
        ]
        counts = _count_exceptions_per_creator(exceptions, MOCK_TT_MAP, MOCK_IG_MAP)
        assert counts["Alice"] == 3


# ===========================================================================
# 6. Pipeline integration — exception counts wired to CreatorSummary
# ===========================================================================

class TestPipelineExceptionWiring:
    """Verify exception counts are correctly wired into CreatorSummary via the pipeline."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_exception_count_in_response(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """
        Bob has 1 unpaired TT → match_exceptions includes Bob's video.
        api_exceptions has 1 private_user (unmappable).
        Bob's CreatorSummary.exception_count should be 1 (only Bob's unpaired).
        """
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        mock_fetch_videos.return_value = (MOCK_VIDEOS, MOCK_API_EXCEPTIONS)

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })
        assert response.status_code == 200

        # Verify the xlsx was generated and we can read it
        data = response.json()
        filepath = os.path.join(output_dir, data["filename"])
        assert os.path.exists(filepath)

        # Load and verify Tab 1 has correct exception counts
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]

        # Find rows for Alice and Bob
        creators_data = {}
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            exc_count = ws.cell(row=row, column=6).value
            creators_data[name] = exc_count

        # Alice: 0 exceptions (her pair matched fine)
        assert creators_data.get("Alice") == 0
        # Bob: 1 exception (his TT was unpaired)
        assert creators_data.get("Bob") == 1


# ===========================================================================
# 7. Multiple creators — full pipeline
# ===========================================================================

class TestMultiCreatorPipeline:
    """Full pipeline with more complex multi-creator data."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_three_creators_various_scenarios(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        """
        Creator A: 2TT + 2IG (all paired, lengths match)
        Creator B: 1TT only (unpaired)
        Creator C: 1IG only (unpaired)
        + 1 unmapped TT video
        + 1 API exception (private)
        """
        creators = [
            Creator(creator_name="CreatorA", tiktok_handle="ca_tt", instagram_handle="ca_ig"),
            Creator(creator_name="CreatorB", tiktok_handle="cb_tt", instagram_handle="cb_ig"),
            Creator(creator_name="CreatorC", tiktok_handle="cc_tt", instagram_handle="cc_ig"),
        ]
        tt_map = {"ca_tt": "CreatorA", "cb_tt": "CreatorB", "cc_tt": "CreatorC"}
        ig_map = {"ca_ig": "CreatorA", "cb_ig": "CreatorB", "cc_ig": "CreatorC"}

        videos = [
            # CreatorA: 2 pairs
            make_video("ca_tt", "tiktok", 30, 50000, "2026-02-20T10:00:00+00:00", "tt_ca1"),
            make_video("ca_ig", "instagram", 30, 80000, "2026-02-20T10:30:00+00:00", "ig_ca1"),
            make_video("ca_tt", "tiktok", 45, 10000, "2026-02-20T14:00:00+00:00", "tt_ca2"),
            make_video("ca_ig", "instagram", 45, 15000, "2026-02-20T14:30:00+00:00", "ig_ca2"),
            # CreatorB: 1 TT only
            make_video("cb_tt", "tiktok", 60, 5000, "2026-02-20T16:00:00+00:00", "tt_cb1"),
            # CreatorC: 1 IG only
            make_video("cc_ig", "instagram", 25, 2500, "2026-02-20T18:00:00+00:00", "ig_cc1"),
            # Unmapped user
            make_video("random_user", "tiktok", 20, 999, "2026-02-20T20:00:00+00:00", "tt_rnd"),
        ]

        api_exceptions = [
            make_exception("ca_tt", "tiktok", "video marked private"),
        ]

        mock_fetch_mapping.return_value = (creators, tt_map, ig_map)
        mock_fetch_videos.return_value = (videos, api_exceptions)

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        assert response.status_code == 200
        data = response.json()
        summary = data["summary"]

        assert summary["total_creators"] == 3
        assert summary["total_videos_processed"] == 7
        # CreatorA: 2 pairs, CreatorB: 1 unpaired, CreatorC: 1 unpaired
        assert summary["total_paired"] == 2
        assert summary["total_unpaired"] == 2

        # Payouts: CreatorA pair1: max(50K,80K)=80K→$100, pair2: max(10K,15K)=15K→$50
        # CreatorB: 5K→$35, CreatorC: 2.5K→$35
        assert summary["total_payout"] == 220.0  # $100 + $50 + $35 + $35

        # Exceptions:
        # 1 api exception (private ca_tt)
        # 1 match: unmapped random_user ("not in creator list")
        # 2 match: unpaired (cb_tt, cc_ig)
        assert summary["total_exceptions"] == 4


# ===========================================================================
# 8. Edge case: all videos are exceptions
# ===========================================================================

class TestAllExceptions:
    """All videos fail filtering → empty payout, only exceptions."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_all_filtered_out(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = (MOCK_CREATORS, MOCK_TT_MAP, MOCK_IG_MAP)
        # All videos are exceptions, no valid videos
        mock_fetch_videos.return_value = ([], [
            make_exception("user1", "tiktok", "video marked private"),
            make_exception("user2", "instagram", "video removed"),
        ])

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        assert response.status_code == 200
        data = response.json()
        assert data["summary"]["total_creators"] == 0
        assert data["summary"]["total_payout"] == 0
        assert data["summary"]["total_paired"] == 0
        assert data["summary"]["total_unpaired"] == 0
        assert data["summary"]["total_exceptions"] == 2


# ===========================================================================
# 9. Edge case: videos but no creator mapping
# ===========================================================================

class TestNoCreatorMapping:
    """Videos exist but no creators in mapping → all go to exceptions."""

    @patch("main.fetch_creator_mapping")
    @patch("main.fetch_videos")
    def test_all_unmapped(self, mock_fetch_videos, mock_fetch_mapping, client, output_dir):
        mock_fetch_mapping.return_value = ([], {}, {})  # No creators
        mock_fetch_videos.return_value = (MOCK_VIDEOS, [])

        response = client.post("/api/calculate", json={
            "start_date": "2026-02-20",
            "end_date": "2026-02-21",
        })

        assert response.status_code == 200
        data = response.json()
        assert data["summary"]["total_creators"] == 0
        assert data["summary"]["total_payout"] == 0
        # All 3 videos should be in exceptions as "not in creator list"
        assert data["summary"]["total_exceptions"] == 3
