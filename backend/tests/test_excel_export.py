"""
Comprehensive tests for services/excel_export.py.

Tests verify:
  1. FILE GENERATION: file created, correct name, correct path
  2. TAB STRUCTURE: 3 tabs with correct names
  3. TAB 1 — Creator Payout Summary:
     - Correct headers, row count, sort order (payout desc), data accuracy
  4. TAB 2 — Video Audit:
     - Correct headers, row count, sort order (name then date)
     - Paired rows have both platforms, unpaired rows have only one
  5. TAB 3 — Exceptions:
     - Correct headers, row count, data accuracy
  6. FORMATTING:
     - Bold header rows, frozen top row, currency + number formats, auto-fit
  7. EDGE CASES:
     - Empty inputs, single row, large datasets, None values
"""

import sys
import os
import pytest
import tempfile
import shutil
from datetime import date, datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from openpyxl.styles import Font

from models.schemas import Video, PayoutUnit, CreatorSummary, ExceptionVideo
from services.excel_export import (
    generate_report,
    _get_uploaded_at,
    _get_video_length,
    _get_latest_updated_at,
    _tab2_sort_key,
    _format_date,
    _format_datetime,
    CURRENCY_FORMAT,
    NUMBER_FORMAT,
)


# ===========================================================================
# Test fixtures
# ===========================================================================

@pytest.fixture
def output_dir():
    """Create a temp directory for output, clean up after test."""
    d = tempfile.mkdtemp(prefix="payout_test_")
    yield d
    shutil.rmtree(d, ignore_errors=True)


def make_video(
    username="test_tt", platform="tiktok", length=30, views=5000,
    uploaded_at_date=None, created_at_dt=None, updated_at_dt=None,
    ad_link=None,
):
    """Helper to create a Video with sensible defaults."""
    return Video(
        username=username,
        platform=platform,
        ad_link=ad_link or f"https://{platform}.com/@{username}/video/123",
        uploaded_at=uploaded_at_date or date(2026, 2, 20),
        created_at=created_at_dt or datetime(2026, 2, 20, 10, 0, 0),
        video_length=length,
        latest_views=views,
        latest_updated_at=updated_at_dt or datetime(2026, 2, 21, 12, 0, 0),
    )


def make_paired_unit(
    creator="Alice", tt_views=50000, ig_views=80000, payout=100.0,
    uploaded_at_date=None, length=30,
):
    """Helper to create a paired PayoutUnit."""
    tt = make_video("alice_tt", "tiktok", length, tt_views,
                    uploaded_at_date=uploaded_at_date,
                    ad_link=f"https://tiktok.com/@alice_tt/{tt_views}")
    ig = make_video("alice_ig", "instagram", length, ig_views,
                    uploaded_at_date=uploaded_at_date,
                    ad_link=f"https://instagram.com/@alice_ig/{ig_views}")
    return PayoutUnit(
        creator_name=creator,
        tiktok_video=tt, instagram_video=ig,
        chosen_views=max(tt_views, ig_views),
        effective_views=min(max(tt_views, ig_views), 10_000_000),
        best_platform="instagram" if ig_views > tt_views else "tiktok",
        payout_amount=payout, paired=True,
        match_confidence="high", pair_note="exact match",
    )


def make_unpaired_unit(
    creator="Bob", platform="tiktok", views=5000, payout=35.0,
    uploaded_at_date=None, length=30,
):
    """Helper to create an unpaired PayoutUnit."""
    video = make_video(f"{creator.lower()}_tt", platform, length, views,
                       uploaded_at_date=uploaded_at_date,
                       ad_link=f"https://{platform}.com/@{creator.lower()}/{views}")
    return PayoutUnit(
        creator_name=creator,
        tiktok_video=video if platform == "tiktok" else None,
        instagram_video=video if platform == "instagram" else None,
        chosen_views=views, effective_views=min(views, 10_000_000),
        best_platform=platform,
        payout_amount=payout, paired=False,
        match_confidence="low", pair_note="unpaired — single platform only",
    )


def make_summary(name="Alice", qualified=2, payout=550.0, paired=1, unpaired=1, exceptions=0):
    """Helper to create a CreatorSummary."""
    return CreatorSummary(
        creator_name=name,
        qualified_video_count=qualified,
        total_payout=payout,
        paired_video_count=paired,
        unpaired_video_count=unpaired,
        exception_count=exceptions,
    )


def make_exception(username="unknown", platform="tiktok", reason="not in creator list",
                   views=1000, length=45):
    """Helper to create an ExceptionVideo."""
    return ExceptionVideo(
        username=username, platform=platform,
        ad_link=f"https://{platform}.com/@{username}/video/456",
        created_at=datetime(2026, 2, 20, 15, 0, 0),
        latest_views=views, video_length=length,
        reason=reason,
    )


# ===========================================================================
# 1. FILE GENERATION
# ===========================================================================

class TestFileGeneration:
    """Verify file is created with correct name and path."""

    def test_file_created(self, output_dir):
        filepath = generate_report(
            [make_summary()], [make_paired_unit()], [make_exception()],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        assert os.path.exists(filepath)

    def test_correct_filename(self, output_dir):
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        expected_name = "Polymarket Payout Summary 2026-02-20 to 2026-02-21.xlsx"
        assert os.path.basename(filepath) == expected_name

    def test_different_date_range(self, output_dir):
        filepath = generate_report(
            [], [], [], date(2026, 1, 1), date(2026, 1, 31), output_dir,
        )
        expected_name = "Polymarket Payout Summary 2026-01-01 to 2026-01-31.xlsx"
        assert os.path.basename(filepath) == expected_name

    def test_output_dir_created(self):
        """Output directory should be auto-created if it doesn't exist."""
        temp_base = tempfile.mkdtemp()
        nested_dir = os.path.join(temp_base, "nested", "deep")
        try:
            filepath = generate_report(
                [], [], [], date(2026, 2, 20), date(2026, 2, 21), nested_dir,
            )
            assert os.path.exists(nested_dir)
            assert os.path.exists(filepath)
        finally:
            shutil.rmtree(temp_base, ignore_errors=True)

    def test_returns_absolute_path(self, output_dir):
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        # filepath should be under the output_dir
        assert filepath.startswith(output_dir)


# ===========================================================================
# 2. TAB STRUCTURE
# ===========================================================================

class TestTabStructure:
    """Verify the workbook has exactly 3 tabs with correct names."""

    def test_three_tabs(self, output_dir):
        filepath = generate_report(
            [make_summary()], [make_paired_unit()], [make_exception()],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        assert len(wb.sheetnames) == 3

    def test_tab_names(self, output_dir):
        filepath = generate_report(
            [make_summary()], [make_paired_unit()], [make_exception()],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        assert wb.sheetnames == [
            "Creator Payout Summary",
            "Video Audit",
            "Exceptions",
        ]


# ===========================================================================
# 3. TAB 1 — Creator Payout Summary
# ===========================================================================

class TestTab1CreatorSummary:
    """Verify Tab 1 data, headers, sort order, and row count."""

    def test_headers(self, output_dir):
        filepath = generate_report(
            [make_summary()], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Creator Name", "Qualified Video Count", "Total Payout",
            "Paired Video Count", "Unpaired Video Count", "Exception Count",
        ]

    def test_row_count(self, output_dir):
        """3 summaries → 1 header + 3 data rows = 4 total rows."""
        summaries = [
            make_summary("Alice", payout=550.0),
            make_summary("Bob", payout=35.0),
            make_summary("Charlie", payout=1000.0),
        ]
        filepath = generate_report(
            summaries, [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        assert ws.max_row == 4  # 1 header + 3 data

    def test_sorted_by_payout_descending(self, output_dir):
        """Rows should be sorted by Total Payout descending."""
        summaries = [
            make_summary("Alice", payout=550.0),
            make_summary("Bob", payout=35.0),
            make_summary("Charlie", payout=1000.0),
        ]
        filepath = generate_report(
            summaries, [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        # Row 2 = highest payout (Charlie, $1000)
        assert ws.cell(row=2, column=1).value == "Charlie"
        assert ws.cell(row=2, column=3).value == 1000.0
        # Row 3 = middle (Alice, $550)
        assert ws.cell(row=3, column=1).value == "Alice"
        # Row 4 = lowest (Bob, $35)
        assert ws.cell(row=4, column=1).value == "Bob"

    def test_data_accuracy(self, output_dir):
        summary = make_summary("TestCreator", qualified=3, payout=650.0,
                               paired=2, unpaired=1, exceptions=1)
        filepath = generate_report(
            [summary], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        # Row 2 is the data row
        assert ws.cell(row=2, column=1).value == "TestCreator"
        assert ws.cell(row=2, column=2).value == 3      # qualified
        assert ws.cell(row=2, column=3).value == 650.0   # payout
        assert ws.cell(row=2, column=4).value == 2       # paired
        assert ws.cell(row=2, column=5).value == 1       # unpaired
        assert ws.cell(row=2, column=6).value == 1       # exceptions

    def test_empty_summaries(self, output_dir):
        """No summaries → just header row."""
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        assert ws.max_row == 1  # header only


# ===========================================================================
# 4. TAB 2 — Video Audit
# ===========================================================================

class TestTab2VideoAudit:
    """Verify Tab 2 data, headers, paired/unpaired distinction, and sorting."""

    def test_headers(self, output_dir):
        filepath = generate_report(
            [], [make_paired_unit()], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Creator Name", "Uploaded At", "Video Length (sec)",
            "TikTok Link", "TikTok Views", "Instagram Link", "Instagram Views",
            "Chosen Views", "Effective Views", "Payout Amount",
            "Paired/Unpaired", "Match Confidence", "Match Notes",
            "Latest Updated At",
        ]

    def test_paired_row_has_both_platforms(self, output_dir):
        """Paired row should have both TikTok and Instagram link + views."""
        unit = make_paired_unit("Alice", tt_views=50000, ig_views=80000)
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        # Row 2 = data row
        assert ws.cell(row=2, column=4).value is not None   # TikTok Link
        assert ws.cell(row=2, column=5).value == 50000       # TikTok Views
        assert ws.cell(row=2, column=6).value is not None   # Instagram Link
        assert ws.cell(row=2, column=7).value == 80000       # Instagram Views
        assert ws.cell(row=2, column=11).value == "paired"

    def test_unpaired_tiktok_row(self, output_dir):
        """Unpaired TikTok row → TikTok columns filled, Instagram blank."""
        unit = make_unpaired_unit("Bob", "tiktok", views=5000)
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.cell(row=2, column=4).value is not None   # TikTok Link
        assert ws.cell(row=2, column=5).value == 5000       # TikTok Views
        assert ws.cell(row=2, column=6).value is None       # Instagram Link (blank)
        assert ws.cell(row=2, column=7).value is None       # Instagram Views (blank)
        assert ws.cell(row=2, column=11).value == "unpaired"

    def test_unpaired_instagram_row(self, output_dir):
        """Unpaired Instagram row → Instagram columns filled, TikTok blank."""
        unit = make_unpaired_unit("Carol", "instagram", views=8000)
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.cell(row=2, column=4).value is None       # TikTok Link (blank)
        assert ws.cell(row=2, column=5).value is None       # TikTok Views (blank)
        assert ws.cell(row=2, column=6).value is not None   # Instagram Link
        assert ws.cell(row=2, column=7).value == 8000       # Instagram Views
        assert ws.cell(row=2, column=11).value == "unpaired"

    def test_sorted_by_creator_then_date(self, output_dir):
        """Tab 2 sorted by Creator Name asc, then Uploaded At asc."""
        units = [
            make_paired_unit("Charlie", uploaded_at_date=date(2026, 2, 21)),
            make_paired_unit("Alice", uploaded_at_date=date(2026, 2, 22)),
            make_paired_unit("Alice", uploaded_at_date=date(2026, 2, 20)),
            make_paired_unit("Bob", uploaded_at_date=date(2026, 2, 20)),
        ]
        filepath = generate_report(
            [], units, [], date(2026, 2, 20), date(2026, 2, 22), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        # Row 2-3: Alice (2020 then 2022)
        assert ws.cell(row=2, column=1).value == "Alice"
        assert ws.cell(row=2, column=2).value == "2026-02-20"
        assert ws.cell(row=3, column=1).value == "Alice"
        assert ws.cell(row=3, column=2).value == "2026-02-22"
        # Row 4: Bob
        assert ws.cell(row=4, column=1).value == "Bob"
        # Row 5: Charlie
        assert ws.cell(row=5, column=1).value == "Charlie"

    def test_payout_amount_and_views(self, output_dir):
        """Verify chosen_views, effective_views, and payout_amount are written correctly."""
        unit = make_paired_unit("Alice", tt_views=50000, ig_views=80000, payout=100.0)
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.cell(row=2, column=8).value == 80000      # Chosen Views
        assert ws.cell(row=2, column=9).value == 80000       # Effective Views
        assert ws.cell(row=2, column=10).value == 100.0      # Payout Amount

    def test_match_metadata(self, output_dir):
        """Verify match confidence and notes are written."""
        unit = make_paired_unit()
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.cell(row=2, column=12).value == "high"         # confidence
        assert ws.cell(row=2, column=13).value == "exact match"  # notes

    def test_row_count(self, output_dir):
        """3 payout units → 1 header + 3 data rows."""
        units = [make_paired_unit(), make_unpaired_unit(), make_paired_unit("C")]
        filepath = generate_report(
            [], units, [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.max_row == 4

    def test_empty_payout_units(self, output_dir):
        """No payout units → just header row."""
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.max_row == 1


# ===========================================================================
# 5. TAB 3 — Exceptions
# ===========================================================================

class TestTab3Exceptions:
    """Verify Tab 3 data and headers."""

    def test_headers(self, output_dir):
        filepath = generate_report(
            [], [], [make_exception()], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        headers = [cell.value for cell in ws[1]]
        assert headers == [
            "Username", "Platform", "Video Link", "Created At",
            "Latest Views", "Video Length (sec)", "Reason",
        ]

    def test_data_accuracy(self, output_dir):
        exc = make_exception("baduser", "instagram", "video marked private",
                             views=500, length=60)
        filepath = generate_report(
            [], [], [exc], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        assert ws.cell(row=2, column=1).value == "baduser"
        assert ws.cell(row=2, column=2).value == "instagram"
        assert ws.cell(row=2, column=5).value == 500
        assert ws.cell(row=2, column=6).value == 60
        assert ws.cell(row=2, column=7).value == "video marked private"

    def test_multiple_exceptions(self, output_dir):
        """Multiple exception types all appear."""
        exceptions = [
            make_exception("user1", "tiktok", "not in creator list"),
            make_exception("user2", "instagram", "video marked private"),
            make_exception("user3", "tiktok", "unpaired — single platform only"),
            make_exception("user4", "tiktok", "missing video length"),
        ]
        filepath = generate_report(
            [], [], exceptions, date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        assert ws.max_row == 5  # 1 header + 4 data

        # Verify all reason types are present
        reasons = [ws.cell(row=r, column=7).value for r in range(2, 6)]
        assert "not in creator list" in reasons
        assert "video marked private" in reasons
        assert "unpaired — single platform only" in reasons
        assert "missing video length" in reasons

    def test_empty_exceptions(self, output_dir):
        """No exceptions → just header row."""
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        assert ws.max_row == 1

    def test_exception_with_none_fields(self, output_dir):
        """Exception with None created_at / video_length should not crash."""
        exc = ExceptionVideo(
            username="null_user", platform="tiktok",
            ad_link="https://tiktok.com/@null/123",
            created_at=None, latest_views=None, video_length=None,
            reason="missing video length",
        )
        filepath = generate_report(
            [], [], [exc], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        assert ws.cell(row=2, column=1).value == "null_user"
        assert ws.cell(row=2, column=4).value is None  # created_at
        assert ws.cell(row=2, column=5).value is None  # latest_views
        assert ws.cell(row=2, column=6).value is None  # video_length


# ===========================================================================
# 6. FORMATTING
# ===========================================================================

class TestFormatting:
    """Verify formatting: bold headers, freeze, currency/number formats, auto-fit."""

    def test_bold_header_tab1(self, output_dir):
        filepath = generate_report(
            [make_summary()], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_bold_header_tab2(self, output_dir):
        filepath = generate_report(
            [], [make_paired_unit()], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_bold_header_tab3(self, output_dir):
        filepath = generate_report(
            [], [], [make_exception()], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        for cell in ws[1]:
            assert cell.font.bold is True

    def test_frozen_top_row_tab1(self, output_dir):
        filepath = generate_report(
            [make_summary()], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        assert ws.freeze_panes == "A2"

    def test_frozen_top_row_tab2(self, output_dir):
        filepath = generate_report(
            [], [make_paired_unit()], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.freeze_panes == "A2"

    def test_frozen_top_row_tab3(self, output_dir):
        filepath = generate_report(
            [], [], [make_exception()], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        assert ws.freeze_panes == "A2"

    def test_currency_format_tab1(self, output_dir):
        """Tab 1 column C (Total Payout) should have currency format."""
        filepath = generate_report(
            [make_summary(payout=550.0)], [], [],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        payout_cell = ws.cell(row=2, column=3)
        assert payout_cell.number_format == CURRENCY_FORMAT

    def test_currency_format_tab2(self, output_dir):
        """Tab 2 column J (Payout Amount) should have currency format."""
        filepath = generate_report(
            [], [make_paired_unit(payout=100.0)], [],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        payout_cell = ws.cell(row=2, column=10)
        assert payout_cell.number_format == CURRENCY_FORMAT

    def test_number_format_views_tab2(self, output_dir):
        """Tab 2 views columns should have comma-separated number format."""
        filepath = generate_report(
            [], [make_paired_unit(tt_views=50000, ig_views=80000)], [],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        # TikTok Views (col E=5), IG Views (col G=7), Chosen (col H=8), Effective (col I=9)
        for col_idx in [5, 7, 8, 9]:
            cell = ws.cell(row=2, column=col_idx)
            if cell.value is not None:
                assert cell.number_format == NUMBER_FORMAT

    def test_number_format_views_tab3(self, output_dir):
        """Tab 3 Latest Views column should have comma-separated number format."""
        filepath = generate_report(
            [], [], [make_exception(views=50000)],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        views_cell = ws.cell(row=2, column=5)
        assert views_cell.number_format == NUMBER_FORMAT

    def test_auto_fit_columns(self, output_dir):
        """Column widths should be > 0 (auto-fit applied)."""
        filepath = generate_report(
            [make_summary()], [make_paired_unit()], [make_exception()],
            date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_idx in range(1, ws.max_column + 1):
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
                width = ws.column_dimensions[col_letter].width
                assert width is not None and width >= 10, (
                    f"{sheet_name} col {col_letter} width={width}"
                )


# ===========================================================================
# 7. EDGE CASES
# ===========================================================================

class TestEdgeCases:
    """Edge case tests."""

    def test_all_empty(self, output_dir):
        """All empty inputs → valid file with just headers."""
        filepath = generate_report(
            [], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        assert os.path.exists(filepath)
        wb = load_workbook(filepath)
        assert len(wb.sheetnames) == 3
        for ws_name in wb.sheetnames:
            assert wb[ws_name].max_row == 1  # header only

    def test_single_creator_single_video(self, output_dir):
        """Minimal data: 1 creator, 1 video, 0 exceptions."""
        summary = make_summary("Solo", qualified=1, payout=35.0, paired=0, unpaired=1)
        unit = make_unpaired_unit("Solo", "tiktok", 5000, 35.0)
        filepath = generate_report(
            [summary], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        assert wb["Creator Payout Summary"].max_row == 2
        assert wb["Video Audit"].max_row == 2
        assert wb["Exceptions"].max_row == 1

    def test_zero_payout_creator(self, output_dir):
        """Creator with $0 payout still appears in Tab 1."""
        summary = make_summary("ZeroGuy", qualified=0, payout=0.0, paired=0, unpaired=1)
        filepath = generate_report(
            [summary], [], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Creator Payout Summary"]
        assert ws.cell(row=2, column=1).value == "ZeroGuy"
        assert ws.cell(row=2, column=3).value == 0.0

    def test_large_dataset(self, output_dir):
        """50 creators, 100 payout units, 20 exceptions → valid file."""
        summaries = [make_summary(f"Creator_{i:03d}", payout=float(i * 100))
                     for i in range(50)]
        units = [make_paired_unit(f"Creator_{i % 50:03d}") for i in range(100)]
        exceptions = [make_exception(f"user_{i}") for i in range(20)]
        filepath = generate_report(
            summaries, units, exceptions,
            date(2026, 2, 1), date(2026, 2, 28), output_dir,
        )
        wb = load_workbook(filepath)
        assert wb["Creator Payout Summary"].max_row == 51   # 1 header + 50
        assert wb["Video Audit"].max_row == 101              # 1 header + 100
        assert wb["Exceptions"].max_row == 21                # 1 header + 20

    def test_capped_views_shown_correctly(self, output_dir):
        """Video with 12M views → chosen=12M, effective=10M in Tab 2."""
        unit = make_paired_unit("BigViews", tt_views=12_000_000, ig_views=1000,
                                payout=2250.0)
        # Override effective_views since our helper caps it
        unit.chosen_views = 12_000_000
        unit.effective_views = 10_000_000
        filepath = generate_report(
            [], [unit], [], date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Video Audit"]
        assert ws.cell(row=2, column=8).value == 12_000_000  # Chosen Views (uncapped)
        assert ws.cell(row=2, column=9).value == 10_000_000  # Effective Views (capped)

    def test_multiple_exception_reasons(self, output_dir):
        """All exception reason types should be preserved."""
        reasons = [
            "not in creator list",
            "video marked private",
            "video removed",
            "missing video length",
            "missing view data",
            "unpaired — single platform only",
        ]
        exceptions = [make_exception(f"user_{i}", reason=r)
                      for i, r in enumerate(reasons)]
        filepath = generate_report(
            [], [], exceptions, date(2026, 2, 20), date(2026, 2, 21), output_dir,
        )
        wb = load_workbook(filepath)
        ws = wb["Exceptions"]
        found_reasons = [ws.cell(row=r, column=7).value for r in range(2, len(reasons) + 2)]
        for reason in reasons:
            assert reason in found_reasons


# ===========================================================================
# 8. HELPER FUNCTION TESTS
# ===========================================================================

class TestHelpers:
    """Test data extraction and formatting helper functions."""

    def test_get_uploaded_at_paired_prefers_tiktok(self):
        unit = make_paired_unit()
        result = _get_uploaded_at(unit)
        assert result == unit.tiktok_video.uploaded_at

    def test_get_uploaded_at_unpaired_instagram(self):
        unit = make_unpaired_unit(platform="instagram")
        result = _get_uploaded_at(unit)
        assert result == unit.instagram_video.uploaded_at

    def test_get_uploaded_at_none_videos(self):
        unit = PayoutUnit(creator_name="X", chosen_views=0)
        result = _get_uploaded_at(unit)
        assert result is None

    def test_get_video_length_paired(self):
        unit = make_paired_unit(length=45)
        assert _get_video_length(unit) == 45

    def test_get_latest_updated_at_picks_max(self):
        tt = make_video("tt", "tiktok", updated_at_dt=datetime(2026, 2, 20, 10, 0))
        ig = make_video("ig", "instagram", updated_at_dt=datetime(2026, 2, 21, 14, 0))
        unit = PayoutUnit(
            creator_name="Test", tiktok_video=tt, instagram_video=ig,
            chosen_views=5000, paired=True,
        )
        result = _get_latest_updated_at(unit)
        assert result == datetime(2026, 2, 21, 14, 0)

    def test_format_date(self):
        assert _format_date(date(2026, 2, 20)) == "2026-02-20"
        assert _format_date(None) is None

    def test_format_datetime(self):
        dt = datetime(2026, 2, 20, 14, 30, 45)
        assert _format_datetime(dt) == "2026-02-20 14:30:45"
        assert _format_datetime(None) is None

    def test_tab2_sort_key_ordering(self):
        """Sort key should order by creator name, then uploaded_at."""
        u1 = make_paired_unit("Bob", uploaded_at_date=date(2026, 2, 22))
        u2 = make_paired_unit("Alice", uploaded_at_date=date(2026, 2, 21))
        u3 = make_paired_unit("Alice", uploaded_at_date=date(2026, 2, 20))

        sorted_units = sorted([u1, u2, u3], key=_tab2_sort_key)
        assert sorted_units[0].creator_name == "Alice"
        assert _get_uploaded_at(sorted_units[0]) == date(2026, 2, 20)
        assert sorted_units[1].creator_name == "Alice"
        assert sorted_units[2].creator_name == "Bob"


# ===========================================================================
# 9. FULL END-TO-END WITH REALISTIC DATA
# ===========================================================================

class TestFullEndToEnd:
    """
    Realistic end-to-end: 2 creators, mixed paired/unpaired, exceptions.
    Verify complete file is valid and data is correctly placed across all tabs.
    """

    def test_realistic_report(self, output_dir):
        # Creator A: $100 + $500 = $600, 2 paired, 0 unpaired, 0 exceptions
        # Creator B: $35, 0 paired, 1 unpaired, 1 exception
        summaries = [
            make_summary("Creator A", qualified=2, payout=600.0, paired=2,
                         unpaired=0, exceptions=0),
            make_summary("Creator B", qualified=1, payout=35.0, paired=0,
                         unpaired=1, exceptions=1),
        ]

        units = [
            make_paired_unit("Creator A", tt_views=50000, ig_views=80000, payout=100.0,
                             uploaded_at_date=date(2026, 2, 20)),
            make_paired_unit("Creator A", tt_views=800000, ig_views=300000, payout=500.0,
                             uploaded_at_date=date(2026, 2, 21)),
            make_unpaired_unit("Creator B", "tiktok", views=5000, payout=35.0,
                               uploaded_at_date=date(2026, 2, 20)),
        ]

        exceptions = [
            make_exception("unknown_user", "tiktok", "not in creator list"),
            make_exception("creator_b_tt", "tiktok", "unpaired — single platform only",
                           views=5000),
        ]

        filepath = generate_report(
            summaries, units, exceptions,
            date(2026, 2, 20), date(2026, 2, 22), output_dir,
        )

        wb = load_workbook(filepath)

        # --- Tab 1: sorted by payout desc ---
        ws1 = wb["Creator Payout Summary"]
        assert ws1.max_row == 3  # header + 2 creators
        # Creator A ($600) should be first
        assert ws1.cell(row=2, column=1).value == "Creator A"
        assert ws1.cell(row=2, column=3).value == 600.0
        # Creator B ($35) should be second
        assert ws1.cell(row=3, column=1).value == "Creator B"
        assert ws1.cell(row=3, column=3).value == 35.0

        # --- Tab 2: sorted by name then date ---
        ws2 = wb["Video Audit"]
        assert ws2.max_row == 4  # header + 3 units
        # Row 2-3: Creator A (sorted by date)
        assert ws2.cell(row=2, column=1).value == "Creator A"
        assert ws2.cell(row=3, column=1).value == "Creator A"
        # Row 4: Creator B
        assert ws2.cell(row=4, column=1).value == "Creator B"

        # --- Tab 3: both exceptions present ---
        ws3 = wb["Exceptions"]
        assert ws3.max_row == 3  # header + 2 exceptions
